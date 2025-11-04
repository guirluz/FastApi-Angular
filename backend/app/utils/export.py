from fpdf import FPDF
from openpyxl import Workbook
from typing import List, Dict
import os
from datetime import datetime

EXPORT_DIR = "/app/uploads"
os.makedirs(EXPORT_DIR, exist_ok=True)


def generate_users_pdf(users: List[Dict]) -> str:
    """
    Genera un archivo PDF con una tabla simple de usuarios.
    Columnas: ID, Username, Email.
    """
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Título
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Reporte de Usuarios", ln=True, align="C")

    # Fecha
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"Fecha: {datetime.utcnow().isoformat()} UTC", ln=True, align="R")

    pdf.ln(5)

    # Encabezados
    pdf.set_font("Arial", "B", 11)
    pdf.cell(20, 8, "ID", border=1, align="C")
    pdf.cell(70, 8, "Usuario", border=1, align="L")
    pdf.cell(100, 8, "Email", border=1, align="L")
    pdf.ln(8)

    # Filas
    pdf.set_font("Arial", "", 10)
    for u in users:
        pdf.cell(20, 8, str(u.get("id", "")), border=1, align="C")
        pdf.cell(70, 8, str(u.get("username", ""))[:35], border=1, align="L")
        pdf.cell(100, 8, str(u.get("email", ""))[:50], border=1, align="L")
        pdf.ln(8)

    filename = f"users_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(EXPORT_DIR, filename)
    pdf.output(filepath)
    return filepath


def generate_users_excel(users: List[Dict]) -> str:
    """
    Genera un archivo Excel con la tabla cruda de usuarios.
    Columnas: id, username, email.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Encabezados
    headers = ["id", "username", "email"]
    ws.append(headers)

    # Filas
    for u in users:
        ws.append([u.get("id", ""), u.get("username", ""), u.get("email", "")])

    filename = f"users_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath

# =====================================================================

def generate_products_pdf(products: List[Dict]) -> str:
    """
    Genera un archivo PDF con los productos.
    Columnas: ID, Nombre, Descripción, Costo por hora, Fecha de registro.
    """
    pdf = FPDF(orientation="L", unit="mm", format="A4")  # Horizontal para más espacio
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Título
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Reporte de Productos", ln=True, align="C")

    # Fecha
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"Fecha: {datetime.utcnow().isoformat()} UTC", ln=True, align="R")
    pdf.ln(5)

    # Encabezados
    pdf.set_font("Arial", "B", 11)
    headers = ["ID", "Nombre", "Descripción", "Costo/Hora", "Fecha Registro"]
    widths = [15, 50, 90, 30, 60]
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 8, h, border=1, align="C")
    pdf.ln(8)

    # Filas
    pdf.set_font("Arial", "", 10)
    for p in products:
        pdf.cell(widths[0], 8, str(p.get("id", "")), border=1, align="C")
        pdf.cell(widths[1], 8, str(p.get("nombre", ""))[:25], border=1, align="L")
        pdf.cell(widths[2], 8, str(p.get("descripcion", ""))[:50], border=1, align="L")
        pdf.cell(widths[3], 8, str(p.get("costo_por_hora", "")), border=1, align="C")
        pdf.cell(widths[4], 8, str(p.get("fecha_registro", "")), border=1, align="C")
        pdf.ln(8)

    filename = f"products_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(EXPORT_DIR, filename)
    pdf.output(filepath)
    return filepath


def generate_products_excel(products: List[Dict]) -> str:
    """
    Genera un archivo Excel con los productos.
    Columnas: id, nombre, descripcion, costo_por_hora, fecha_registro.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["id", "nombre", "descripcion", "costo_por_hora", "fecha_registro"]
    ws.append(headers)

    for p in products:
        ws.append([
            p.get("id", ""),
            p.get("nombre", ""),
            p.get("descripcion", ""),
            p.get("costo_por_hora", ""),
            str(p.get("fecha_registro", "")),
        ])

    filename = f"products_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath

# =============================================================

def generate_rentals_pdf(rentals: List[Dict]) -> str:
    """
    Genera un archivo PDF con las rentas.
    Columnas: ID, Usuario, Producto, Horas rentadas, Costo total, Fecha renta.
    """
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Título
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Reporte de Rentas", ln=True, align="C")

    # Fecha
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"Fecha: {datetime.utcnow().isoformat()} UTC", ln=True, align="R")
    pdf.ln(5)

    # Encabezados
    pdf.set_font("Arial", "B", 11)
    headers = ["ID", "Usuario", "Producto", "Horas", "Costo Total", "Fecha Renta"]
    widths = [15, 50, 60, 25, 30, 60]
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 8, h, border=1, align="C")
    pdf.ln(8)

    # Filas
    pdf.set_font("Arial", "", 10)
    for r in rentals:
        pdf.cell(widths[0], 8, str(r.get("id", "")), border=1, align="C")
        pdf.cell(widths[1], 8, str(r.get("usuario", ""))[:25], border=1, align="L")
        pdf.cell(widths[2], 8, str(r.get("producto", ""))[:30], border=1, align="L")
        pdf.cell(widths[3], 8, str(r.get("horas_rentadas", "")), border=1, align="C")
        pdf.cell(widths[4], 8, str(r.get("costo_total", "")), border=1, align="C")
        pdf.cell(widths[5], 8, str(r.get("fecha_renta", "")), border=1, align="C")
        pdf.ln(8)

    filename = f"rentals_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(EXPORT_DIR, filename)
    pdf.output(filepath)
    return filepath


def generate_rentals_excel(rentals: List[Dict]) -> str:
    """
    Genera un archivo Excel con las rentas.
    Columnas: id, usuario, producto, horas_rentadas, costo_total, fecha_renta.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rentas"

    headers = ["id", "usuario", "producto", "horas_rentadas", "costo_total", "fecha_renta"]
    ws.append(headers)

    for r in rentals:
        ws.append([
            r.get("id", ""),
            r.get("usuario", ""),
            r.get("producto", ""),
            r.get("horas_rentadas", ""),
            r.get("costo_total", ""),
            str(r.get("fecha_renta", "")),
        ])

    filename = f"rentals_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath
